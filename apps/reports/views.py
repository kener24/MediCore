from datetime import timedelta
from decimal import Decimal
from io import BytesIO

from django.db.models import Avg, Count, F, Q, Sum
from django.db.models.functions import TruncDate, TruncMonth
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.accounts.models import User
from apps.accounts.permissions import get_role_name
from apps.appointments.models import Appointment
from apps.billing.models import CashMovement, CashSession, Invoice, InvoiceItem, Payment
from apps.clinics.models import Clinic
from apps.doctors.models import DoctorProfile
from apps.inventory.models import InventoryCategory, InventoryItem, InventoryLot, InventoryMovement
from apps.medical_records.models import ClinicalConsultation
from apps.patients.models import Patient
from apps.prescriptions.models import Diagnosis
from apps.purchases.models import PurchaseOrder, PurchaseOrderItem, Supplier
from apps.audit.models import AuditLog
from apps.audit.services import log_audit_event
from apps.reports.serializers import ReportDateFiltersSerializer


ADMIN_REPORT_ROLES = ["admin", "recepcionista", "enfermera"]
FINANCIAL_ROLES = ["admin", "recepcionista", "cajero", "recepcionista_caja"]


def zero():
    return Decimal("0.00")


def role(user):
    return get_role_name(user)


def forbidden(message="No tienes permiso para ver estos reportes."):
    return Response({"detail": message}, status=status.HTTP_403_FORBIDDEN)


def validate_filters(request):
    serializer = ReportDateFiltersSerializer(data=request.query_params)
    serializer.is_valid(raise_exception=True)
    today = timezone.localdate()
    data = serializer.validated_data
    return {
        **data,
        "date_from": data.get("date_from") or today - timedelta(days=30),
        "date_to": data.get("date_to") or today,
    }


def clinic_id_for_request(request, filters):
    user = request.user
    if role(user) == "superadmin" or user.is_superuser:
        return None
    return user.clinica_id


def clinic_scope(request, queryset, filters=None, field="clinic_id"):
    filters = filters or {}
    clinic_id = clinic_id_for_request(request, filters)
    return queryset.filter(**{field: clinic_id}) if clinic_id else queryset


def date_scope(queryset, filters, field):
    return queryset.filter(**{f"{field}__gte": filters["date_from"], f"{field}__lte": filters["date_to"]})


def grouped_count(queryset, field):
    return {row[field] or "sin_dato": row["count"] for row in queryset.values(field).annotate(count=Count("id")).order_by(field)}


def by_day(queryset, date_field, value_field=None):
    qs = queryset.values(day=F(date_field))
    qs = qs.annotate(value=Sum(value_field) if value_field else Count("id")).order_by("day")
    return [{"date": row["day"], "amount" if value_field else "count": row["value"] or zero()} for row in qs]


def by_month(queryset, date_field, value_field=None):
    qs = queryset.annotate(month=TruncMonth(date_field)).values("month")
    qs = qs.annotate(value=Sum(value_field) if value_field else Count("id")).order_by("month")
    return [{"month": row["month"].date() if hasattr(row["month"], "date") else row["month"], "amount" if value_field else "count": row["value"] or zero()} for row in qs]


def as_text(value):
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return float(value)
    return value


def report_filename(report, extension):
    stamp = timezone.localtime().strftime("%Y%m%d-%H%M")
    return f"medicore-{report}-{stamp}.{extension}"


def report_scope(request, filters):
    return clinic_id_for_request(request, filters)


def purchase_received_amount(order):
    return sum((item.quantity_received * item.unit_cost for item in order.items.all()), zero())


def purchase_pending_amount(order):
    return sum((item.pending_quantity * item.unit_cost for item in order.items.all()), zero())


def export_payload(request, report):
    filters = validate_filters(request)
    title = report.replace("-", " ").title()
    generated = timezone.localtime().strftime("%Y-%m-%d %H:%M")
    common = [
        ("Reporte", title),
        ("Fecha desde", filters["date_from"]),
        ("Fecha hasta", filters["date_to"]),
        ("Generado", generated),
        ("Usuario", request.user.email),
    ]

    if report == "clinic-dashboard":
        if role(request.user) not in ADMIN_REPORT_ROLES:
            return None, forbidden()
        patients = clinic_scope(request, Patient.objects.all(), filters)
        appointments = date_scope(clinic_scope(request, Appointment.objects.all(), filters), filters, "scheduled_date")
        consultations = date_scope(clinic_scope(request, ClinicalConsultation.objects.all(), filters), filters, "consultation_date")
        invoices = date_scope(clinic_scope(request, Invoice.objects.all(), filters), filters, "issue_date")
        payments = date_scope(clinic_scope(request, Payment.objects.filter(active=True, status=Payment.Status.APLICADO), filters), filters, "payment_date")
        items = clinic_scope(request, InventoryItem.objects.all(), filters)
        top_doctors = consultations.values("doctor_id", "doctor__user__nombre_completo").annotate(consultations=Count("id")).order_by("-consultations")[:20]
        summary = common + [
            ("Pacientes", patients.count()),
            ("Citas", appointments.count()),
            ("Consultas finalizadas", consultations.filter(status=ClinicalConsultation.Status.FINALIZADA).count()),
            ("Facturado", invoices.aggregate(total=Sum("total_amount"))["total"] or zero()),
            ("Pagado", payments.aggregate(total=Sum("amount"))["total"] or zero()),
            ("Bajo stock", items.filter(stock_current__lte=F("stock_minimum")).count()),
        ]
        headers = ["Medico", "Consultas"]
        rows = [[row["doctor__user__nombre_completo"], row["consultations"]] for row in top_doctors]
        return {"title": title, "summary": summary, "headers": headers, "rows": rows}, None

    if report == "superadmin-dashboard":
        if role(request.user) != "superadmin" and not request.user.is_superuser:
            return None, forbidden()
        appointments = date_scope(Appointment.objects.all(), filters, "scheduled_date")
        invoices = date_scope(Invoice.objects.all(), filters, "issue_date")
        headers = ["Clinica", "Pacientes", "Citas", "Revenue"]
        rows = [
            [
                clinic.nombre,
                Patient.objects.filter(clinic=clinic).count(),
                appointments.filter(clinic=clinic).count(),
                invoices.filter(clinic=clinic).aggregate(total=Sum("total_amount"))["total"] or zero(),
            ]
            for clinic in Clinic.objects.all()[:5000]
        ]
        summary = common + [
            ("Clinicas", Clinic.objects.count()),
            ("Clinicas activas", Clinic.objects.filter(activo=True).count()),
            ("Usuarios", User.objects.count()),
            ("Pacientes", Patient.objects.count()),
            ("Citas periodo", appointments.count()),
            ("Revenue periodo", invoices.aggregate(total=Sum("total_amount"))["total"] or zero()),
        ]
        return {"title": title, "summary": summary, "headers": headers, "rows": rows}, None

    if report == "doctor-dashboard":
        if role(request.user) != "medico":
            return None, forbidden()
        doctor = getattr(request.user, "doctor_profile", None)
        if not doctor:
            return None, Response({"detail": "No tienes perfil medico."}, status=status.HTTP_404_NOT_FOUND)
        appointments = Appointment.objects.filter(doctor=doctor, scheduled_date__gte=filters["date_from"], scheduled_date__lte=filters["date_to"])
        consultations = ClinicalConsultation.objects.filter(doctor=doctor, consultation_date__gte=filters["date_from"], consultation_date__lte=filters["date_to"])
        by_status = appointments.values("status").annotate(count=Count("id")).order_by("status")
        summary = common + [
            ("Citas", appointments.count()),
            ("Consultas finalizadas", consultations.filter(status=ClinicalConsultation.Status.FINALIZADA).count()),
            ("Consultas pendientes", consultations.filter(status=ClinicalConsultation.Status.BORRADOR).count()),
            ("Pacientes atendidos", consultations.values("patient_id").distinct().count()),
        ]
        headers = ["Estado cita", "Cantidad"]
        rows = [[row["status"], row["count"]] for row in by_status]
        return {"title": title, "summary": summary, "headers": headers, "rows": rows}, None

    if report == "reception-dashboard":
        if role(request.user) not in ["admin", "recepcionista"]:
            return None, forbidden()
        today = timezone.localdate()
        appointments_today = clinic_scope(request, Appointment.objects.filter(scheduled_date=today), filters)
        payments_today = clinic_scope(request, Payment.objects.filter(payment_date=today, active=True, status=Payment.Status.APLICADO), filters)
        invoices = clinic_scope(request, Invoice.objects.all(), filters)
        summary = common + [
            ("Citas hoy", appointments_today.count()),
            ("Citas pendientes hoy", appointments_today.filter(status__in=[Appointment.Status.PENDIENTE, Appointment.Status.CONFIRMADA]).count()),
            ("Pacientes registrados hoy", clinic_scope(request, Patient.objects.filter(creado_en__date=today), filters).count()),
            ("Pagos hoy", payments_today.aggregate(total=Sum("amount"))["total"] or zero()),
            ("Facturas pendientes", invoices.filter(status__in=[Invoice.Status.PENDIENTE, Invoice.Status.PARCIAL]).count()),
        ]
        headers = ["Hora", "Paciente", "Medico", "Estado", "Motivo"]
        rows = [
            [a.start_time, a.patient.nombre_completo, a.doctor.user.nombre_completo, a.status, a.reason]
            for a in appointments_today.select_related("patient", "doctor__user").order_by("start_time")[:5000]
        ]
        return {"title": title, "summary": summary, "headers": headers, "rows": rows}, None

    if report == "patients":
        if role(request.user) not in ADMIN_REPORT_ROLES:
            return None, forbidden()
        qs = clinic_scope(request, Patient.objects.all(), filters)
        if request.query_params.get("gender"):
            qs = qs.filter(genero=request.query_params["gender"])
        if request.query_params.get("status"):
            qs = qs.filter(activo=request.query_params["status"] in ["activo", "true", "1"])
        if request.query_params.get("search"):
            s = request.query_params["search"]
            qs = qs.filter(Q(nombres__icontains=s) | Q(apellidos__icontains=s) | Q(identidad__icontains=s) | Q(codigo_paciente__icontains=s))
        headers = ["Codigo", "Nombre", "Identidad", "Genero", "Telefono", "Correo", "Registro", "Estado"]
        rows = [[p.codigo_paciente, p.nombre_completo, p.identidad, p.genero, p.telefono, p.correo, p.creado_en.date(), "Activo" if p.activo else "Inactivo"] for p in qs[:5000]]
        summary = common + [("Total pacientes", qs.count()), ("Activos", qs.filter(activo=True).count()), ("Inactivos", qs.filter(activo=False).count())]
        return {"title": title, "summary": summary, "headers": headers, "rows": rows}, None

    if report == "doctors":
        if role(request.user) not in ADMIN_REPORT_ROLES:
            return None, forbidden()
        qs = clinic_scope(request, DoctorProfile.objects.select_related("user", "specialty"), filters)
        if filters.get("doctor"):
            qs = qs.filter(id=filters["doctor"])
        if request.query_params.get("specialty"):
            qs = qs.filter(specialty_id=request.query_params["specialty"])
        if request.query_params.get("status"):
            qs = qs.filter(activo=request.query_params["status"] in ["activo", "true", "1"])
        consultation_counts = dict(
            date_scope(clinic_scope(request, ClinicalConsultation.objects.all(), filters), filters, "consultation_date")
            .values_list("doctor_id")
            .annotate(count=Count("id"))
        )
        appointment_counts = dict(
            date_scope(clinic_scope(request, Appointment.objects.all(), filters), filters, "scheduled_date")
            .values_list("doctor_id")
            .annotate(count=Count("id"))
        )
        headers = ["Medico", "Especialidad", "Colegiacion", "Tarifa", "Duracion", "Consultas", "Citas", "Estado"]
        rows = [[d.user.nombre_completo, d.specialty.nombre, d.numero_colegiacion, d.tarifa_consulta, d.duracion_consulta_minutos, consultation_counts.get(d.id, 0), appointment_counts.get(d.id, 0), "Activo" if d.activo else "Inactivo"] for d in qs[:5000]]
        summary = common + [("Total medicos", qs.count()), ("Activos", qs.filter(activo=True).count())]
        return {"title": title, "summary": summary, "headers": headers, "rows": rows}, None

    if report == "appointments":
        if role(request.user) not in ADMIN_REPORT_ROLES + ["medico"]:
            return None, forbidden()
        qs = date_scope(clinic_scope(request, Appointment.objects.select_related("patient", "doctor__user", "doctor__specialty", "created_by"), filters), filters, "scheduled_date")
        if role(request.user) == "medico":
            qs = qs.filter(doctor__user=request.user)
        for key in ["doctor", "patient", "status"]:
            if filters.get(key):
                qs = qs.filter(**{f"{key}_id" if key != "status" else key: filters[key]})
        headers = ["Fecha", "Hora", "Paciente", "Medico", "Especialidad", "Estado", "Motivo", "Creada por"]
        rows = [[a.scheduled_date, a.start_time, a.patient.nombre_completo, a.doctor.user.nombre_completo, getattr(a.doctor.specialty, "nombre", ""), a.status, a.reason, getattr(a.created_by, "nombre_completo", "")] for a in qs[:5000]]
        summary = common + [("Total citas", qs.count()), ("Canceladas", qs.filter(status=Appointment.Status.CANCELADA).count()), ("No asistio", qs.filter(status=Appointment.Status.NO_ASISTIO).count())]
        return {"title": title, "summary": summary, "headers": headers, "rows": rows}, None

    if report == "consultations":
        if role(request.user) not in ADMIN_REPORT_ROLES + ["medico"]:
            return None, forbidden()
        qs = date_scope(
            clinic_scope(request, ClinicalConsultation.objects.select_related("patient", "doctor__user", "medical_record"), filters),
            filters,
            "consultation_date",
        )
        if role(request.user) == "medico":
            qs = qs.filter(doctor__user=request.user)
        if filters.get("doctor"):
            qs = qs.filter(doctor_id=filters["doctor"])
        if filters.get("patient"):
            qs = qs.filter(patient_id=filters["patient"])
        if filters.get("status"):
            qs = qs.filter(status=filters["status"])
        headers = ["Fecha", "Paciente", "Medico", "Expediente", "Estado", "Motivo", "Diagnostico", "Finalizada"]
        rows = [[c.consultation_date, c.patient.nombre_completo, c.doctor.user.nombre_completo, c.medical_record.record_number, c.status, c.chief_complaint, c.preliminary_diagnosis, c.finalized_at] for c in qs[:5000]]
        summary = common + [("Total consultas", qs.count()), ("Finalizadas", qs.filter(status=ClinicalConsultation.Status.FINALIZADA).count()), ("Borrador", qs.filter(status=ClinicalConsultation.Status.BORRADOR).count())]
        return {"title": title, "summary": summary, "headers": headers, "rows": rows}, None

    if report == "financial":
        if role(request.user) not in FINANCIAL_ROLES:
            return None, forbidden()
        qs = date_scope(clinic_scope(request, Invoice.objects.select_related("patient"), filters), filters, "issue_date")
        if request.query_params.get("status"):
            qs = qs.filter(status=request.query_params["status"])
        headers = ["Factura", "Fecha", "Paciente", "Subtotal", "Descuento", "Impuesto", "Total", "Pagado", "Saldo", "Estado"]
        rows = [[i.invoice_number, i.issue_date, i.patient.nombre_completo, i.subtotal, i.discount_amount, i.tax_amount, i.total_amount, i.paid_amount, i.balance_due, i.status] for i in qs[:5000]]
        summary = common + [("Total facturado", qs.aggregate(v=Sum("total_amount"))["v"] or zero()), ("Pagado", qs.aggregate(v=Sum("paid_amount"))["v"] or zero()), ("Pendiente", qs.aggregate(v=Sum("balance_due"))["v"] or zero())]
        return {"title": title, "summary": summary, "headers": headers, "rows": rows}, None

    if report == "payments":
        if role(request.user) not in FINANCIAL_ROLES:
            return None, forbidden()
        qs = date_scope(clinic_scope(request, Payment.objects.select_related("invoice", "patient", "received_by"), filters), filters, "payment_date")
        if request.query_params.get("method"):
            qs = qs.filter(method=request.query_params["method"])
        headers = ["Pago", "Fecha", "Factura", "Paciente", "Metodo", "Referencia", "Monto", "Recibido por", "Estado"]
        rows = [[p.payment_number, p.payment_date, p.invoice.invoice_number, p.patient.nombre_completo, p.method, p.reference, p.amount, getattr(p.received_by, "nombre_completo", ""), p.status] for p in qs[:5000]]
        summary = common + [("Total pagos", qs.aggregate(v=Sum("amount"))["v"] or zero()), ("Registros", qs.count())]
        return {"title": title, "summary": summary, "headers": headers, "rows": rows}, None

    if report == "cash":
        if role(request.user) not in FINANCIAL_ROLES:
            return None, forbidden()
        qs = clinic_scope(request, CashSession.objects.select_related("opened_by"), filters).filter(opening_datetime__date__gte=filters["date_from"], opening_datetime__date__lte=filters["date_to"])
        headers = ["Caja", "Apertura", "Cierre", "Usuario", "Monto apertura", "Esperado", "Cierre monto", "Diferencia", "Estado"]
        rows = [[s.id, s.opening_datetime, s.closing_datetime, s.opened_by.nombre_completo, s.opening_amount, s.expected_amount, s.closing_amount, s.difference_amount, s.status] for s in qs[:5000]]
        summary = common + [("Sesiones", qs.count()), ("Abiertas", qs.filter(status=CashSession.Status.ABIERTA).count()), ("Cerradas", qs.filter(status=CashSession.Status.CERRADA).count())]
        return {"title": title, "summary": summary, "headers": headers, "rows": rows}, None

    if report == "inventory":
        if role(request.user) not in ADMIN_REPORT_ROLES:
            return None, forbidden()
        qs = clinic_scope(request, InventoryItem.objects.select_related("category"), filters)
        if request.query_params.get("category"):
            qs = qs.filter(category_id=request.query_params["category"])
        if request.query_params.get("item_type"):
            qs = qs.filter(item_type=request.query_params["item_type"])
        if request.query_params.get("low_stock") == "true":
            qs = qs.filter(stock_current__lte=F("stock_minimum"))
        headers = ["Producto", "SKU", "Categoria", "Tipo", "Stock", "Minimo", "Costo", "Venta", "Valor stock", "Estado"]
        rows = [[i.name, i.sku, getattr(i.category, "name", ""), i.item_type, i.stock_current, i.stock_minimum, i.cost_price, i.sale_price, i.stock_current * i.cost_price, "Activo" if i.active else "Inactivo"] for i in qs[:5000]]
        summary = common + [("Productos", qs.count()), ("Bajo stock", qs.filter(stock_current__lte=F("stock_minimum")).count()), ("Valor stock", qs.aggregate(v=Sum(F("stock_current") * F("cost_price")))["v"] or zero())]
        return {"title": title, "summary": summary, "headers": headers, "rows": rows}, None

    if report == "purchases":
        if role(request.user) != "admin":
            return None, forbidden()
        qs = date_scope(clinic_scope(request, PurchaseOrder.objects.select_related("supplier").prefetch_related("items"), filters), filters, "order_date")
        if request.query_params.get("supplier"):
            qs = qs.filter(supplier_id=request.query_params["supplier"])
        if request.query_params.get("status"):
            qs = qs.filter(status=request.query_params["status"])
        headers = ["Orden", "Fecha", "Proveedor", "Estado", "Subtotal", "Impuesto", "Total", "Recibido", "Pendiente"]
        rows = [[o.order_number, o.order_date, o.supplier.name, o.status, o.subtotal, o.tax_amount, o.total_amount, purchase_received_amount(o), purchase_pending_amount(o)] for o in qs[:5000]]
        summary = common + [("Ordenes", qs.count()), ("Total comprado", qs.aggregate(v=Sum("total_amount"))["v"] or zero())]
        return {"title": title, "summary": summary, "headers": headers, "rows": rows}, None

    if report == "audit":
        if role(request.user) not in ["superadmin", "admin"]:
            return None, forbidden()
        qs = clinic_scope(request, AuditLog.objects.select_related("user", "clinic"), filters).filter(created_at__date__gte=filters["date_from"], created_at__date__lte=filters["date_to"])
        for param in ["action", "module", "severity"]:
            if request.query_params.get(param):
                qs = qs.filter(**{param: request.query_params[param]})
        headers = ["Fecha", "Usuario", "Clinica", "Accion", "Modulo", "Descripcion", "Severidad", "IP"]
        rows = [[l.created_at, getattr(l.user, "email", ""), getattr(l.clinic, "nombre", ""), l.action, l.module, l.description, l.severity, l.ip_address] for l in qs[:5000]]
        summary = common + [("Eventos", qs.count()), ("Criticos", qs.filter(severity=AuditLog.Severity.CRITICAL).count()), ("Errores", qs.filter(severity=AuditLog.Severity.ERROR).count())]
        return {"title": title, "summary": summary, "headers": headers, "rows": rows}, None

    return None, Response({"detail": "Reporte no soportado."}, status=status.HTTP_404_NOT_FOUND)


class ReportExcelExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, report):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        payload, error = export_payload(request, report)
        if error:
            return error
        wb = Workbook()
        ws = wb.active
        ws.title = payload["title"][:31]
        ws.append([payload["title"]])
        ws["A1"].font = Font(bold=True, size=14)
        ws.append([])
        for label, value in payload["summary"]:
            ws.append([label, as_text(value)])
        ws.append([])
        ws.append(payload["headers"])
        header_row = ws.max_row
        for cell in ws[header_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1D4ED8")
        for row in payload["rows"]:
            ws.append([as_text(value) for value in row])
        for index, column in enumerate(ws.columns, start=1):
            width = max(len(str(cell.value or "")) for cell in column)
            ws.column_dimensions[get_column_letter(index)].width = min(max(width + 2, 12), 40)
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        response = HttpResponse(stream.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="{report_filename(report, "xlsx")}"'
        log_audit_event(request=request, action=AuditLog.Action.EXPORT, module=AuditLog.Module.REPORTS, model_name="Report", object_id=report, object_repr=report, description="Reporte exportado a Excel.", new_values={"report": report, "format": "xlsx"})
        return response


class ReportPdfExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, report):
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import landscape, letter
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table as PdfTable, TableStyle

        payload, error = export_payload(request, report)
        if error:
            return error
        stream = BytesIO()
        doc = SimpleDocTemplate(stream, pagesize=landscape(letter), rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)
        styles = getSampleStyleSheet()
        story = [Paragraph(payload["title"], styles["Title"]), Spacer(1, 10)]
        for label, value in payload["summary"]:
            story.append(Paragraph(f"<b>{label}:</b> {as_text(value)}", styles["Normal"]))
        story.append(Spacer(1, 12))
        data = [payload["headers"]] + [[str(as_text(value)) for value in row] for row in payload["rows"][:200]]
        table = PdfTable(data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1D4ED8")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
        ]))
        story.append(table)
        doc.build(story)
        response = HttpResponse(stream.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{report_filename(report, "pdf")}"'
        log_audit_event(request=request, action=AuditLog.Action.EXPORT, module=AuditLog.Module.REPORTS, model_name="Report", object_id=report, object_repr=report, description="Reporte exportado a PDF.", new_values={"report": report, "format": "pdf"})
        return response


class ClinicDashboardReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if role(request.user) not in ADMIN_REPORT_ROLES:
            return forbidden()
        filters = validate_filters(request)
        patients = clinic_scope(request, Patient.objects.all(), filters)
        appointments = date_scope(clinic_scope(request, Appointment.objects.all(), filters), filters, "scheduled_date")
        consultations = date_scope(clinic_scope(request, ClinicalConsultation.objects.all(), filters), filters, "consultation_date")
        invoices = date_scope(clinic_scope(request, Invoice.objects.all(), filters), filters, "issue_date")
        payments = date_scope(clinic_scope(request, Payment.objects.filter(active=True, status=Payment.Status.APLICADO), filters), filters, "payment_date")
        items = clinic_scope(request, InventoryItem.objects.all(), filters)
        today = timezone.localdate()
        top_doctors = consultations.values("doctor_id", "doctor__user__nombre_completo").annotate(consultations=Count("id")).order_by("-consultations")[:5]
        return Response({
            "summary": {
                "total_patients": patients.count(),
                "new_patients": patients.filter(creado_en__date__gte=filters["date_from"], creado_en__date__lte=filters["date_to"]).count(),
                "total_appointments": appointments.count(),
                "appointments_today": clinic_scope(request, Appointment.objects.filter(scheduled_date=today), filters).count(),
                "completed_consultations": consultations.filter(status=ClinicalConsultation.Status.FINALIZADA).count(),
                "pending_appointments": appointments.filter(status__in=[Appointment.Status.PENDIENTE, Appointment.Status.CONFIRMADA]).count(),
                "total_invoiced": invoices.aggregate(total=Sum("total_amount"))["total"] or zero(),
                "total_paid": payments.aggregate(total=Sum("amount"))["total"] or zero(),
                "pending_balance": invoices.aggregate(total=Sum("balance_due"))["total"] or zero(),
                "low_stock_items": items.filter(stock_current__lte=F("stock_minimum")).count(),
            },
            "appointments_by_status": grouped_count(appointments, "status"),
            "revenue_by_day": by_day(payments, "payment_date", "amount"),
            "top_doctors": [{"doctor_id": row["doctor_id"], "doctor_name": row["doctor__user__nombre_completo"], "consultations": row["consultations"]} for row in top_doctors],
        })


class SuperAdminDashboardReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if role(request.user) != "superadmin" and not request.user.is_superuser:
            return forbidden()
        filters = validate_filters(request)
        appointments = date_scope(Appointment.objects.all(), filters, "scheduled_date")
        invoices = date_scope(Invoice.objects.all(), filters, "issue_date")
        overview = []
        for clinic in Clinic.objects.all():
            overview.append({
                "clinic_id": clinic.id,
                "clinic_name": clinic.nombre,
                "patients": Patient.objects.filter(clinic=clinic).count(),
                "appointments": appointments.filter(clinic=clinic).count(),
                "revenue": invoices.filter(clinic=clinic).aggregate(total=Sum("total_amount"))["total"] or zero(),
            })
        return Response({
            "summary": {
                "total_clinics": Clinic.objects.count(),
                "active_clinics": Clinic.objects.filter(activo=True).count(),
                "inactive_clinics": Clinic.objects.filter(activo=False).count(),
                "total_users": User.objects.count(),
                "total_patients": Patient.objects.count(),
                "total_appointments": appointments.count(),
                "total_revenue": invoices.aggregate(total=Sum("total_amount"))["total"] or zero(),
            },
            "clinics_overview": overview,
        })


class AppointmentsReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if role(request.user) not in ADMIN_REPORT_ROLES + ["medico"]:
            return forbidden()
        filters = validate_filters(request)
        qs = date_scope(clinic_scope(request, Appointment.objects.all(), filters), filters, "scheduled_date")
        if role(request.user) == "medico":
            qs = qs.filter(doctor__user=request.user)
        if filters.get("doctor"):
            qs = qs.filter(doctor_id=filters["doctor"])
        if filters.get("patient"):
            qs = qs.filter(patient_id=filters["patient"])
        if filters.get("status"):
            qs = qs.filter(status=filters["status"])
        total = qs.count()
        cancelled = qs.filter(status=Appointment.Status.CANCELADA).count()
        no_show = qs.filter(status=Appointment.Status.NO_ASISTIO).count()
        by_doctor = qs.values("doctor_id", "doctor__user__nombre_completo").annotate(count=Count("id")).order_by("-count")
        return Response({
            "total_appointments": total,
            "appointments_by_status": grouped_count(qs, "status"),
            "appointments_by_day": by_day(qs, "scheduled_date"),
            "appointments_by_doctor": [{"doctor_id": r["doctor_id"], "doctor_name": r["doctor__user__nombre_completo"], "count": r["count"]} for r in by_doctor],
            "cancellation_rate": round((cancelled / total) * 100, 2) if total else 0,
            "no_show_rate": round((no_show / total) * 100, 2) if total else 0,
        })


class PatientsReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if role(request.user) not in ADMIN_REPORT_ROLES:
            return forbidden()
        filters = validate_filters(request)
        qs = clinic_scope(request, Patient.objects.all(), filters)
        created = qs.filter(creado_en__date__gte=filters["date_from"], creado_en__date__lte=filters["date_to"])
        if request.query_params.get("gender"):
            qs = qs.filter(genero=request.query_params["gender"])
            created = created.filter(genero=request.query_params["gender"])
        return Response({
            "total_patients": qs.count(),
            "new_patients": created.count(),
            "active_patients": qs.filter(activo=True).count(),
            "inactive_patients": qs.filter(activo=False).count(),
            "patients_by_gender": grouped_count(qs, "genero"),
            "patients_by_age_range": {},
            "patients_by_month": by_month(created, "creado_en"),
        })


class DoctorsReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if role(request.user) not in ADMIN_REPORT_ROLES:
            return forbidden()
        filters = validate_filters(request)
        doctors = clinic_scope(request, DoctorProfile.objects.all(), filters)
        consultations = date_scope(clinic_scope(request, ClinicalConsultation.objects.all(), filters), filters, "consultation_date")
        appointments = date_scope(clinic_scope(request, Appointment.objects.all(), filters), filters, "scheduled_date")
        if filters.get("doctor"):
            doctors = doctors.filter(id=filters["doctor"])
            consultations = consultations.filter(doctor_id=filters["doctor"])
            appointments = appointments.filter(doctor_id=filters["doctor"])
        if request.query_params.get("specialty"):
            doctors = doctors.filter(specialty_id=request.query_params["specialty"])
        c_by_doc = consultations.values("doctor_id", "doctor__user__nombre_completo").annotate(count=Count("id")).order_by("-count")
        a_by_doc = appointments.values("doctor_id", "doctor__user__nombre_completo").annotate(count=Count("id")).order_by("-count")
        return Response({
            "total_doctors": doctors.count(),
            "active_doctors": doctors.filter(activo=True).count(),
            "consultations_by_doctor": [{"doctor_id": r["doctor_id"], "doctor_name": r["doctor__user__nombre_completo"], "count": r["count"]} for r in c_by_doc],
            "appointments_by_doctor": [{"doctor_id": r["doctor_id"], "doctor_name": r["doctor__user__nombre_completo"], "count": r["count"]} for r in a_by_doc],
            "revenue_by_doctor": [],
            "average_consultations_per_doctor": round((consultations.count() / doctors.count()), 2) if doctors.count() else 0,
        })


class ConsultationsReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if role(request.user) not in ADMIN_REPORT_ROLES + ["medico"]:
            return forbidden()
        filters = validate_filters(request)
        qs = date_scope(clinic_scope(request, ClinicalConsultation.objects.all(), filters), filters, "consultation_date")
        if role(request.user) == "medico":
            qs = qs.filter(doctor__user=request.user)
        if filters.get("doctor"):
            qs = qs.filter(doctor_id=filters["doctor"])
        if filters.get("patient"):
            qs = qs.filter(patient_id=filters["patient"])
        if filters.get("status"):
            qs = qs.filter(status=filters["status"])
        by_doc = qs.values("doctor_id", "doctor__user__nombre_completo").annotate(count=Count("id")).order_by("-count")
        diagnoses = Diagnosis.objects.filter(consultation__in=qs, activo=True).values("name").annotate(count=Count("id")).order_by("-count")[:10]
        return Response({
            "total_consultations": qs.count(),
            "draft_consultations": qs.filter(status=ClinicalConsultation.Status.BORRADOR).count(),
            "finalized_consultations": qs.filter(status=ClinicalConsultation.Status.FINALIZADA).count(),
            "consultations_by_day": by_day(qs, "consultation_date"),
            "consultations_by_doctor": [{"doctor_id": r["doctor_id"], "doctor_name": r["doctor__user__nombre_completo"], "count": r["count"]} for r in by_doc],
            "most_common_diagnoses": [{"name": r["name"], "count": r["count"]} for r in diagnoses],
        })


class FinancialReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if role(request.user) not in FINANCIAL_ROLES:
            return forbidden()
        filters = validate_filters(request)
        invoices = date_scope(clinic_scope(request, Invoice.objects.all(), filters), filters, "issue_date")
        payments = date_scope(clinic_scope(request, Payment.objects.filter(active=True, status=Payment.Status.APLICADO), filters), filters, "payment_date")
        if request.query_params.get("status"):
            invoices = invoices.filter(status=request.query_params["status"])
        if request.query_params.get("payment_method"):
            payments = payments.filter(method=request.query_params["payment_method"])
        top_services = InvoiceItem.objects.filter(invoice__in=invoices, active=True).values("description").annotate(total=Sum("line_total"), count=Count("id")).order_by("-total")[:10]
        count = invoices.count()
        total_invoiced = invoices.aggregate(total=Sum("total_amount"))["total"] or zero()
        return Response({
            "total_invoiced": total_invoiced,
            "total_paid": payments.aggregate(total=Sum("amount"))["total"] or zero(),
            "total_pending": invoices.aggregate(total=Sum("balance_due"))["total"] or zero(),
            "invoices_by_status": grouped_count(invoices, "status"),
            "payments_by_method": grouped_count(payments, "method"),
            "revenue_by_day": by_day(payments, "payment_date", "amount"),
            "revenue_by_month": by_month(payments, "payment_date", "amount"),
            "top_services": [{"description": r["description"], "total": r["total"] or zero(), "count": r["count"]} for r in top_services],
            "average_ticket": total_invoiced / count if count else zero(),
        })


class CashReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if role(request.user) not in FINANCIAL_ROLES:
            return forbidden()
        filters = validate_filters(request)
        sessions = clinic_scope(request, CashSession.objects.all(), filters).filter(opening_datetime__date__gte=filters["date_from"], opening_datetime__date__lte=filters["date_to"])
        movements = clinic_scope(request, CashMovement.objects.filter(active=True), filters).filter(creado_en__date__gte=filters["date_from"], creado_en__date__lte=filters["date_to"])
        payments = date_scope(clinic_scope(request, Payment.objects.filter(active=True, status=Payment.Status.APLICADO, method=Payment.Method.EFECTIVO), filters), filters, "payment_date")
        if request.query_params.get("user"):
            sessions = sessions.filter(opened_by_id=request.query_params["user"])
        summary = sessions.values("opened_by_id", "opened_by__nombre_completo").annotate(count=Count("id"), difference=Sum("difference_amount")).order_by("-count")
        return Response({
            "total_cash_sessions": sessions.count(),
            "open_cash_sessions": sessions.filter(status=CashSession.Status.ABIERTA).count(),
            "closed_cash_sessions": sessions.filter(status=CashSession.Status.CERRADA).count(),
            "cash_received": payments.aggregate(total=Sum("amount"))["total"] or zero(),
            "cash_movements_income": movements.filter(movement_type=CashMovement.Type.INGRESO).aggregate(total=Sum("amount"))["total"] or zero(),
            "cash_movements_expense": movements.filter(movement_type=CashMovement.Type.EGRESO).aggregate(total=Sum("amount"))["total"] or zero(),
            "differences_total": sessions.aggregate(total=Sum("difference_amount"))["total"] or zero(),
            "sessions_summary": [{"user_id": r["opened_by_id"], "user_name": r["opened_by__nombre_completo"], "count": r["count"], "difference": r["difference"] or zero()} for r in summary],
        })


class InventoryReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if role(request.user) not in ADMIN_REPORT_ROLES:
            return forbidden()
        filters = validate_filters(request)
        items = clinic_scope(request, InventoryItem.objects.all(), filters)
        lots = clinic_scope(request, InventoryLot.objects.all(), filters)
        movements = clinic_scope(request, InventoryMovement.objects.all(), filters)
        if request.query_params.get("item_type"):
            items = items.filter(item_type=request.query_params["item_type"])
        if request.query_params.get("category"):
            items = items.filter(category_id=request.query_params["category"])
        today = timezone.localdate()
        moved = movements.values("item_id", "item__name").annotate(quantity=Sum("quantity"), count=Count("id")).order_by("-count")[:10]
        return Response({
            "total_items": items.count(),
            "active_items": items.filter(active=True).count(),
            "low_stock_items": items.filter(stock_current__lte=F("stock_minimum")).count(),
            "expired_lots": lots.filter(expiration_date__lt=today).count(),
            "expiring_soon_lots": lots.filter(expiration_date__gte=today, expiration_date__lte=today + timedelta(days=30)).count(),
            "total_stock_value": items.aggregate(total=Sum(F("stock_current") * F("cost_price")))["total"] or zero(),
            "stock_by_category": [{"category_id": r["category_id"], "category_name": r["category__name"], "stock": r["stock"] or zero()} for r in items.values("category_id", "category__name").annotate(stock=Sum("stock_current")).order_by("category__name")],
            "stock_by_type": [{"type": r["item_type"], "stock": r["stock"] or zero()} for r in items.values("item_type").annotate(stock=Sum("stock_current")).order_by("item_type")],
            "most_moved_items": [{"item_id": r["item_id"], "item_name": r["item__name"], "quantity": r["quantity"] or zero(), "count": r["count"]} for r in moved],
        })


class PurchasesReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if role(request.user) != "admin":
            return forbidden()
        filters = validate_filters(request)
        orders = date_scope(clinic_scope(request, PurchaseOrder.objects.all(), filters), filters, "order_date")
        if request.query_params.get("supplier"):
            orders = orders.filter(supplier_id=request.query_params["supplier"])
        if request.query_params.get("status"):
            orders = orders.filter(status=request.query_params["status"])
        by_supplier = orders.values("supplier_id", "supplier__name").annotate(total=Sum("total_amount"), count=Count("id")).order_by("-total")
        top_items = PurchaseOrderItem.objects.filter(purchase_order__in=orders, active=True).values("item_id", "item__name").annotate(quantity=Sum("quantity_ordered"), total=Sum("line_total")).order_by("-quantity")[:10]
        return Response({
            "total_purchase_orders": orders.count(),
            "purchases_by_status": grouped_count(orders, "status"),
            "total_purchased_amount": orders.exclude(status=PurchaseOrder.Status.CANCELADA).aggregate(total=Sum("total_amount"))["total"] or zero(),
            "purchases_by_supplier": [{"supplier_id": r["supplier_id"], "supplier_name": r["supplier__name"], "total": r["total"] or zero(), "count": r["count"]} for r in by_supplier],
            "purchases_by_month": by_month(orders, "order_date", "total_amount"),
            "pending_receipts": orders.filter(status__in=[PurchaseOrder.Status.PENDIENTE, PurchaseOrder.Status.APROBADA, PurchaseOrder.Status.RECIBIDA_PARCIAL]).count(),
            "top_purchased_items": [{"item_id": r["item_id"], "item_name": r["item__name"], "quantity": r["quantity"] or zero(), "total": r["total"] or zero()} for r in top_items],
        })


class DoctorDashboardReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if role(request.user) != "medico":
            return forbidden()
        filters = validate_filters(request)
        doctor = getattr(request.user, "doctor_profile", None)
        if not doctor:
            return Response({"detail": "No tienes perfil medico."}, status=status.HTTP_404_NOT_FOUND)
        today = timezone.localdate()
        appointments = Appointment.objects.filter(doctor=doctor, scheduled_date__gte=filters["date_from"], scheduled_date__lte=filters["date_to"])
        consultations = ClinicalConsultation.objects.filter(doctor=doctor, consultation_date__gte=filters["date_from"], consultation_date__lte=filters["date_to"])
        return Response({
            "summary": {
                "appointments_today": Appointment.objects.filter(doctor=doctor, scheduled_date=today).count(),
                "upcoming_appointments": Appointment.objects.filter(doctor=doctor, scheduled_date__gte=today, status__in=[Appointment.Status.PENDIENTE, Appointment.Status.CONFIRMADA]).count(),
                "completed_consultations": consultations.filter(status=ClinicalConsultation.Status.FINALIZADA).count(),
                "pending_consultations": consultations.filter(status=ClinicalConsultation.Status.BORRADOR).count(),
                "patients_attended": consultations.values("patient_id").distinct().count(),
            },
            "appointments_by_status": grouped_count(appointments, "status"),
            "consultations_by_day": by_day(consultations, "consultation_date"),
        })


class ReceptionDashboardReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if role(request.user) not in ["admin", "recepcionista"]:
            return forbidden()
        filters = validate_filters(request)
        today = timezone.localdate()
        appointments_today = clinic_scope(request, Appointment.objects.filter(scheduled_date=today), filters)
        payments_today = clinic_scope(request, Payment.objects.filter(payment_date=today, active=True, status=Payment.Status.APLICADO), filters)
        invoices = clinic_scope(request, Invoice.objects.all(), filters)
        current_cash = clinic_scope(request, CashSession.objects.filter(opened_by=request.user, status=CashSession.Status.ABIERTA), filters).first()
        return Response({
            "appointments_today": appointments_today.count(),
            "pending_appointments": appointments_today.filter(status__in=[Appointment.Status.PENDIENTE, Appointment.Status.CONFIRMADA]).count(),
            "patients_registered_today": clinic_scope(request, Patient.objects.filter(creado_en__date=today), filters).count(),
            "payments_today": payments_today.aggregate(total=Sum("amount"))["total"] or zero(),
            "current_cash": {"id": current_cash.id, "opening_amount": current_cash.opening_amount, "expected_amount": current_cash.expected_amount} if current_cash else None,
            "pending_invoices": invoices.filter(status__in=[Invoice.Status.PENDIENTE, Invoice.Status.PARCIAL]).count(),
        })
