from io import BytesIO

from django.db.models import Count, Q
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.pagination import PageNumberPagination
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from apps.accounts.permissions import get_role_name
from apps.audit.models import AuditLog
from apps.audit.serializers import AuditFilterSerializer, AuditLogDetailSerializer, AuditLogListSerializer, AuditLogStatsSerializer
from apps.audit.services import create_audit_log


def can_view_audit(user):
    return bool(user and user.is_authenticated and (user.is_superuser or get_role_name(user) in ["superadmin", "admin"]))


def is_super(user):
    return bool(user and (user.is_superuser or get_role_name(user) == "superadmin"))


def scoped_logs(request):
    queryset = AuditLog.objects.select_related("clinic", "user", "user__role")
    if is_super(request.user):
        clinic = request.query_params.get("clinic")
        if clinic:
            queryset = queryset.filter(clinic_id=clinic)
    elif get_role_name(request.user) == "admin" and request.user.clinica_id:
        queryset = queryset.filter(clinic_id=request.user.clinica_id)
    else:
        queryset = queryset.none()
    return queryset


def apply_filters(request, queryset):
    filters = AuditFilterSerializer(data=request.query_params)
    filters.is_valid(raise_exception=True)
    params = filters.validated_data

    if params.get("user"):
        queryset = queryset.filter(user_id=params["user"])
    for param in ["action", "module", "severity", "status", "object_id"]:
        if params.get(param):
            queryset = queryset.filter(**{param: params[param]})
    if params.get("object_type"):
        queryset = queryset.filter(Q(object_type=params["object_type"]) | Q(model_name=params["object_type"]))
    if params.get("model_name"):
        queryset = queryset.filter(Q(object_type=params["model_name"]) | Q(model_name=params["model_name"]))
    if params.get("date_from"):
        queryset = queryset.filter(created_at__date__gte=params["date_from"])
    if params.get("date_to"):
        queryset = queryset.filter(created_at__date__lte=params["date_to"])
    if params.get("search"):
        search = params["search"]
        queryset = queryset.filter(
            Q(description__icontains=search)
            | Q(object_repr__icontains=search)
            | Q(request_path__icontains=search)
            | Q(user_email__icontains=search)
            | Q(user__email__icontains=search)
        )
    return queryset


class AuditPagination(PageNumberPagination):
    page_size = 50
    page_size_query_param = "page_size"
    max_page_size = 200


def _audit_export_rows(queryset):
    rows = []
    for log in queryset[:5000]:
        rows.append(
            [
                timezone.localtime(log.created_at).strftime("%Y-%m-%d %H:%M:%S"),
                log.user_email or getattr(log.user, "email", ""),
                log.user_role or getattr(getattr(log.user, "role", None), "nombre", ""),
                getattr(log.clinic, "nombre", ""),
                log.module,
                log.action,
                log.description,
                log.severity,
                log.status,
                log.ip_address or "",
                log.request_method,
                log.request_path,
            ]
        )
    return rows


class AuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    permission_classes = [IsAuthenticated]
    queryset = AuditLog.objects.select_related("clinic", "user", "user__role")
    pagination_class = AuditPagination

    def get_serializer_class(self):
        return AuditLogDetailSerializer if self.action == "retrieve" else AuditLogListSerializer

    def get_queryset(self):
        if not can_view_audit(self.request.user):
            return AuditLog.objects.none()
        return apply_filters(self.request, scoped_logs(self.request))

    def list(self, request, *args, **kwargs):
        if not can_view_audit(request.user):
            create_audit_log(
                request=request,
                action=AuditLog.Action.PERMISSION_DENIED,
                module=AuditLog.Module.SECURITY,
                description="Intento no autorizado de acceso a auditoria.",
                status=AuditLog.Status.FAILED,
                severity=AuditLog.Severity.WARNING,
            )
            return Response({"detail": "No tienes permiso para ver la bitacora."}, status=status.HTTP_403_FORBIDDEN)
        return super().list(request, *args, **kwargs)

    def retrieve(self, request, *args, **kwargs):
        if not can_view_audit(request.user):
            return Response({"detail": "No tienes permiso para ver la bitacora."}, status=status.HTTP_403_FORBIDDEN)
        response = super().retrieve(request, *args, **kwargs)
        if response.status_code == status.HTTP_200_OK:
            create_audit_log(
                request=request,
                action=AuditLog.Action.VIEW,
                module=AuditLog.Module.SECURITY,
                obj=self.get_object(),
                object_type="AuditLog",
                description="Detalle de auditoria consultado.",
            )
        return response

    @action(detail=False, methods=["get"], url_path="my-activity")
    def my_activity(self, request):
        queryset = AuditLog.objects.select_related("clinic", "user", "user__role").filter(user=request.user)
        queryset = apply_filters(request, queryset)
        page = self.paginate_queryset(queryset)
        serializer = AuditLogListSerializer(page if page is not None else queryset, many=True)
        return self.get_paginated_response(serializer.data) if page is not None else Response(serializer.data)

    @action(detail=False, methods=["get"], url_path="export-excel")
    def export_excel(self, request):
        if not can_view_audit(request.user):
            return Response({"detail": "No tienes permiso para exportar auditoria."}, status=status.HTTP_403_FORBIDDEN)
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        queryset = self.get_queryset()
        headers = ["Fecha", "Usuario", "Rol", "Clinica", "Modulo", "Accion", "Descripcion", "Severidad", "Estado", "IP", "Metodo", "Ruta"]
        wb = Workbook()
        ws = wb.active
        ws.title = "Auditoria"
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0F766E")
        for row in _audit_export_rows(queryset):
            ws.append(row)
        for idx, _ in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = 22
        output = BytesIO()
        wb.save(output)
        create_audit_log(request=request, action=AuditLog.Action.EXPORT, module=AuditLog.Module.SECURITY, description="Auditoria exportada a Excel.")
        response = HttpResponse(output.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="auditoria.xlsx"'
        return response

    @action(detail=False, methods=["get"], url_path="export-pdf")
    def export_pdf(self, request):
        if not can_view_audit(request.user):
            return Response({"detail": "No tienes permiso para exportar auditoria."}, status=status.HTTP_403_FORBIDDEN)
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import landscape, letter
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table as PdfTable, TableStyle

        queryset = self.get_queryset()
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(letter), leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
        styles = getSampleStyleSheet()
        data = [["Fecha", "Usuario", "Clinica", "Modulo", "Accion", "Severidad", "Estado", "IP"]]
        for row in _audit_export_rows(queryset)[:1000]:
            data.append([row[0], row[1], row[3], row[4], row[5], row[7], row[8], row[9]])
        table = PdfTable(data, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F766E")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                ]
            )
        )
        doc.build([Paragraph("Auditoria MediCore", styles["Title"]), Spacer(1, 12), table])
        create_audit_log(request=request, action=AuditLog.Action.EXPORT, module=AuditLog.Module.SECURITY, description="Auditoria exportada a PDF.")
        response = HttpResponse(output.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="auditoria.pdf"'
        return response


class AuditStatsViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]

    def list(self, request):
        if not can_view_audit(request.user):
            return Response({"detail": "No tienes permiso para ver estadisticas de auditoria."}, status=status.HTTP_403_FORBIDDEN)
        queryset = apply_filters(request, scoped_logs(request))
        today = timezone.localdate()
        data = {
            "total_logs": queryset.count(),
            "logs_today": queryset.filter(created_at__date=today).count(),
            "warnings": queryset.filter(severity=AuditLog.Severity.WARNING).count(),
            "errors": queryset.filter(severity=AuditLog.Severity.ERROR).count(),
            "critical": queryset.filter(severity=AuditLog.Severity.CRITICAL).count(),
            "failed": queryset.filter(status=AuditLog.Status.FAILED).count(),
            "login_success": queryset.filter(action=AuditLog.Action.LOGIN_SUCCESS).count(),
            "login_failed": queryset.filter(action=AuditLog.Action.LOGIN_FAILED).count(),
            "top_actions": list(queryset.values("action").annotate(count=Count("id")).order_by("-count")[:8]),
            "top_modules": list(queryset.values("module").annotate(count=Count("id")).order_by("-count")[:8]),
        }
        return Response(AuditLogStatsSerializer(data).data)
